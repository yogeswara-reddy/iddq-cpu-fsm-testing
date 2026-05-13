#!/usr/bin/env python3
"""Generate RWU-RV64I Cache / AMAT / Clock-Matrix Excel workbook."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "RWU_RV64I_Cache_Analysis.xlsx")

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def fill(hex6): return PatternFill("solid", fgColor=hex6)
def fwh(bold=True, sz=10): return Font(color="FFFFFF", bold=bold, size=sz)
def fbk(bold=False, sz=10, color="000000"): return Font(color=color, bold=bold, size=sz)

FILL_DARK   = fill("1F4E79")   # dark blue  – main header
FILL_MED    = fill("2E75B6")   # mid  blue  – sub-header
FILL_LIGHT  = fill("D6E4F0")   # light blue – section bg
FILL_GREEN  = fill("E2EFDA")   # light green
FILL_DGRN   = fill("375623")   # dark green – best option
FILL_BGRN   = fill("70AD47")   # bright green – good
FILL_YELLOW = fill("FFFF99")   # yellow – marginal
FILL_ORANGE = fill("FFC000")   # orange – warning
FILL_RED    = fill("FFB3B3")   # red – bad
FILL_GREY   = fill("F2F2F2")   # alternating row
FILL_WHITE  = fill("FFFFFF")

thn = Side(border_style="thin",   color="000000")
med = Side(border_style="medium", color="000000")
THIN = Border(left=thn, right=thn, top=thn, bottom=thn)
MED  = Border(left=med, right=med, top=med, bottom=med)
LMED = Border(left=med, right=thn, top=thn, bottom=thn)

def s(ws, row, col, val=None, f=None, fnt=None, ha="left", va="center",
      brd=THIN, nf=None, wrap=False, rs=1, cs=1):
    if rs > 1 or cs > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+rs-1, end_column=col+cs-1)
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    if f:   c.fill  = f
    if fnt: c.font  = fnt
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if brd: c.border = brd
    if nf:  c.number_format = nf
    return c

def hdr(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_DARK, fwh(sz=11), ha="center", cs=cs)

def shdr(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_MED, fwh(sz=10), ha="center", cs=cs)

def sh2(ws, row, col, val, cs=1):
    return s(ws, row, col, val, FILL_LIGHT, fbk(bold=True), cs=cs)

def lbl(ws, row, col, val, bold=True, wrap=False):
    return s(ws, row, col, val, fnt=fbk(bold=bold), wrap=wrap)

def num(ws, row, col, val, nf="0.00", f=None):
    return s(ws, row, col, val, f=f, fnt=fbk(), ha="center", nf=nf)

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

# ===========================================================================
# SHEET 1 – Legende
# ===========================================================================
ws1 = wb.active
ws1.title = "Legende"
for col, w in zip("ABCDE", [30, 28, 48, 22, 22]):
    ws1.column_dimensions[get_column_letter(col.encode()[0]-64)].width = w

r = 1
# Title
s(ws1, r, 1, "RWU-RV64I  —  Cache-Speichersystem  (Hennessy-Patterson Analyse)",
  FILL_DARK, fwh(sz=13), ha="center", cs=5); ws1.row_dimensions[r].height=28; r+=1
s(ws1, r, 1, "Prozess: X-Fab XO035 (350 nm)  |  Architektur: Harvard  |  "
  "Konfiguration: I-Cache + D-Cache (je 4-way SA, 4 KB), Scratchpad SRAM",
  FILL_MED, fwh(sz=10, bold=False), ha="center", cs=5); r+=2

# --- Cache-Eigenschaften -------------------------------------------------
shdr(ws1, r, 1, "Eigenschaft", cs=1)
shdr(ws1, r, 2, "I-Cache", cs=1)
shdr(ws1, r, 3, "D-Cache (Flash-Region)", cs=1)
shdr(ws1, r, 4, "D-Cache (Scratchpad-Region)", cs=1)
shdr(ws1, r, 5, "Scratchpad SRAM (Bypass)", cs=1)
r+=1
rows_cache = [
    ("Kapazität",              "4 KB (param.)",           "4 KB (param.)",           "n/a (bypass)",              "TBD (param.)"),
    ("Assoziativität",         "4-fach set-assoziativ",   "4-fach set-assoziativ",   "—",                         "direkt adressiert"),
    ("Sets / Cacheline",       "32 Sets, 32 B",           "32 Sets, 32 B",           "—",                         "—"),
    ("Ersetzungsstrategie",    "Pseudo-LRU (3 Bit/Set)",  "Pseudo-LRU (3 Bit/Set)",  "—",                         "—"),
    ("Write Policy",           "read-only (kein Write)",  "read-allocate, no-write-back",  "kein Cache (bypass)",  "write-through direkt"),
    ("Write-Allocate",         "nicht anwendbar",         "nein  (Flash read-only)",  "nein (bypass)",            "nicht anwendbar"),
    ("Dirty Bit",              "nein",                    "nein (eliminiert)",        "nein",                     "—"),
    ("Hit-Latenz",             "1 Systemtakt",            "1 Systemtakt",            "1–2 Systemtakte (bypass)",  "1–2 Systemtakte"),
    ("Miss-Penalty",           "106 × f_sys/f_QSPI Takte","106 × f_sys/f_QSPI Takte","—  (kein Miss möglich)",   "—"),
    ("AXI4 Write-Path",        "nein",                    "nein (eliminiert)",        "nein",                     "nein"),
    ("AXI4 Read-Path",         "ja  (AR + R Kanal)",      "ja  (AR + R Kanal)",       "nein",                     "nein"),
    ("Bus zum Speicher",       "AXI4 Burst (ARLEN=3)",    "AXI4 Burst (ARLEN=3)",    "SRAM-Interface (sync.)",    "SRAM-Interface (sync.)"),
]
for prop, ic, dc_fl, dc_sp, sp in rows_cache:
    bg = FILL_GREY if (r % 2 == 0) else FILL_WHITE
    lbl(ws1, r, 1, prop, bold=True)
    for ci, val in enumerate([ic, dc_fl, dc_sp, sp], start=2):
        s(ws1, r, ci, val, f=bg, fnt=fbk(bold=False), wrap=True)
    r+=1
r+=1

# --- H&P Terminologie -------------------------------------------------------
shdr(ws1, r, 1, "H&P Begriff", cs=2)
shdr(ws1, r, 3, "Bedeutung / Formel", cs=2)
shdr(ws1, r, 5, "In diesem System")
r+=1
terms = [
    ("Hit Time  (t_hit)",
     "Zugriffszeit bei Cache-Treffer",
     "1 Systemtakt (I+D)"),
    ("Miss Rate  (m)",
     "Anteil Zugriffe ohne Treffer",
     "I-Cache: abhängig vom Code-Footprint\nD-Cache (Flash): abhängig von .rodata-Zugriffen"),
    ("Miss Penalty  (t_miss)",
     "Mehrzyklen bei Miss bis Daten bereit stehen",
     "106 × (f_sys / f_QSPI_SCK)  Takte\n≡ 4.24 µs bei 25 MHz QSPI (beide Flash-Typen)"),
    ("AMAT",
     "Average Memory Access Time\n= t_hit + m × t_miss",
     "I: 1 + m_I × t_miss\nD: 1 + m_D × t_miss  (nur Load-Misses aus Flash)"),
    ("Memory Stall CPI",
     "Taktzyklen/Instr. für Speicherwartezustände\n= m_I×t_miss + load_freq×m_D×t_miss",
     "Store-Misses = 0 (Scratchpad-Bypass, kein Flash-Write)"),
    ("CPI_eff",
     "Effektiver CPI\n= CPI_ideal + Memory_Stall_CPI",
     "CPI_ideal = 1 (No-Pipeline-CPU)"),
    ("Write-Through",
     "Jeder Write sofort in Backing Store",
     "NICHT verwendet (kein Flash-Write)"),
    ("Write-Back",
     "Write erst bei Eviction in Backing Store",
     "NICHT verwendet (Flash read-only, Dirty-Bit eliminiert)"),
    ("Write-Allocate",
     "Miss bei Write → Cache-Line laden",
     "NICHT verwendet (Stores gehen direkt → Scratchpad)"),
    ("No-Write-Allocate",
     "Miss bei Write → direkt in Backing Store",
     "NICHT relevant (Scratchpad ist nicht Flash)"),
    ("Read-Allocate",
     "Miss bei Load → neue Line in Cache laden",
     "JA – für I-Cache und D-Cache (Flash-Region)"),
    ("Pseudo-LRU (PLRU)",
     "3-Bit-Approximation des LRU für 4-Wege",
     "1 PLRU-Feld pro Set (96 Bit Tag-SRAM enthält 3 PLRU-Bits)"),
    ("Scratchpad",
     "Direkt adressierter SRAM (kein Cache)",
     "Hält Stack, Heap, .data, .bss\nD-Cache-Controller bypassed direkt"),
]
for i, (term, defn, sys_note) in enumerate(terms):
    bg = FILL_GREY if (i % 2 == 0) else FILL_WHITE
    s(ws1, r, 1, term, f=FILL_LIGHT, fnt=fbk(bold=True), cs=2, wrap=True)
    s(ws1, r, 3, defn, f=bg, fnt=fbk(), cs=2, wrap=True)
    s(ws1, r, 5, sys_note, f=bg, fnt=fbk(), wrap=True)
    ws1.row_dimensions[r].height = 32
    r+=1
r+=1

# --- Linker-Segmente --------------------------------------------------------
shdr(ws1, r, 1, "Linker-Segment", cs=1)
shdr(ws1, r, 2, "Read/Write", cs=1)
shdr(ws1, r, 3, "Inhalt", cs=2)
shdr(ws1, r, 5, "Speicherort")
r+=1
segs = [
    (".text",    "RO", "Programmcode (Instruktionen)",                  "NOR Flash → I-Cache"),
    (".rodata",  "RO", "Konstanten, String-Literale",                   "NOR Flash → D-Cache (Flash-Region)"),
    (".data",    "RW", "Initialisierte globale/statische Variablen",    "Scratchpad (Kopie von Flash bei Boot)"),
    (".bss",     "RW", "Uninitial. glob./stat. Var. (Null-Init.)",      "Scratchpad (Start-up nullt Region)"),
    ("Heap",     "RW", "Dynamisch alloz. Speicher (malloc/free)",       "Scratchpad (wächst ↑ von .bss)"),
    ("Stack",    "RW", "Auto-Variablen, Rücksprungadr., Register-Save", "Scratchpad (wächst ↓ von oben)"),
]
for i, (seg, rw, content, where) in enumerate(segs):
    bg = FILL_GREEN if rw == "RO" else FILL_YELLOW
    s(ws1, r, 1, seg,     f=bg,       fnt=fbk(bold=True), ha="center")
    s(ws1, r, 2, rw,      f=bg,       fnt=fbk(bold=False), ha="center")
    s(ws1, r, 3, content, f=FILL_WHITE, fnt=fbk(), cs=2, wrap=True)
    s(ws1, r, 5, where,   f=FILL_WHITE, fnt=fbk())
    r+=1

# ===========================================================================
# SHEET 2 – AMAT Szenarien (Hennessy-Patterson)
# ===========================================================================
ws2 = wb.create_sheet("AMAT_Szenarien")
for col, w in zip(range(1, 14), [6, 22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]):
    ws2.column_dimensions[get_column_letter(col)].width = w

r = 1
s(ws2, r, 1, "AMAT-Analyse nach Hennessy-Patterson  —  RWU-RV64I Cache-System",
  FILL_DARK, fwh(sz=12), ha="center", cs=13)
ws2.row_dimensions[r].height = 24; r+=2

# --- Parameter Box ----------------------------------------------------------
shdr(ws2, r, 1, "System-Parameter", cs=5)
r+=1
params = [
    ("t_hit (I-Cache + D-Cache)",    "1",   "Systemtakt",   "Kombinatorisch: Tag-Vergleich + SRAM-Read in 1 Takt"),
    ("t_hit (Scratchpad Bypass)",    "1–2", "Systemtakte",  "D-Cache-Controller → SRAM direkt"),
    ("QSPI-Protokoll-Zyklen (tot.)", "106", "QSPI-SCK-Takte","8 (Opcode)+24 (Adr)+8 (Dummy)+64 (Daten, Quad)+2 (CS#)"),
    ("Miss-Penalty  t_miss",         "106 × (f_sys / f_QSPI)", "Systemtakte", "Skalar von Taktver­hältnis, siehe Takt-Matrix"),
    ("Beispiel: f_sys=25 MHz, f_QSPI=25 MHz", "106", "Systemtakte", "≡ 4.24 µs"),
    ("Beispiel: f_sys=10 MHz, f_QSPI=25 MHz", " 42", "Systemtakte", "≡ 4.24 µs (Spec-Referenzpunkt)"),
    ("CPI_ideal  (No-Pipeline)",     "1",   "Takte/Instr.", "Ein Fetch pro Takt (ohne Speicherstalls)"),
    ("Instruction Mix (typisch)",    "—",   "—",            "25 % Load, 10 % Store, 20 % Branch, 45 % ALU"),
    ("Store-Miss-Penalty",           "0",   "—",            "Stores → Scratchpad Bypass (kein Flash-Write)"),
]
for i, (name, val, unit, note) in enumerate(params):
    bg = FILL_GREY if i % 2 == 0 else FILL_WHITE
    s(ws2, r, 1, name,  f=FILL_LIGHT, fnt=fbk(bold=True), cs=2, wrap=True)
    s(ws2, r, 3, val,   f=bg, fnt=fbk(bold=True), ha="center")
    s(ws2, r, 4, unit,  f=bg, fnt=fbk())
    s(ws2, r, 5, note,  f=bg, fnt=fbk(), wrap=True)
    r+=1
r+=1

# --- AMAT-Formel Box --------------------------------------------------------
shdr(ws2, r, 1, "Hennessy-Patterson Formeln (angewendet auf RWU-RV64I)", cs=13); r+=1
formeln = [
    "AMAT  =  t_hit  +  m  ×  t_miss",
    "CPI_eff  =  CPI_ideal  +  I$_stall  +  D$_stall",
    "I$_stall  =  m_I  ×  t_miss                          [Takte pro Instruktion]",
    "D$_stall  =  load_freq × m_D × t_miss                [Store-Miss = 0, Scratchpad-Bypass!]",
    "Memory_Stall_CPI  =  m_I × t_miss  +  0.25 × m_D × t_miss",
]
for f_str in formeln:
    s(ws2, r, 1, f_str, FILL_LIGHT, fbk(bold=True, sz=10), cs=13)
    ws2.row_dimensions[r].height = 18; r+=1
r+=1

# --- AMAT-Tabellen pro t_miss -----------------------------------------------
MISS_PENALTIES = [
    (42,  "f_sys=10 MHz, f_QSPI=25 MHz  (Spec-Referenz)"),
    (106, "f_sys=25 MHz, f_QSPI=25 MHz  (Empfehlung: kein PLL)"),
    (212, "f_sys=50 MHz, f_QSPI=25 MHz  (Empfehlung: PLL×2)"),
    (106, "f_sys=50 MHz, f_QSPI=50 MHz  (aggressiv, div=1)"),
]
MISS_RATES = [0.001, 0.005, 0.01, 0.02, 0.05, 0.10, 0.15, 0.20, 0.30, 0.50]
COL_LABELS = ["Miss-Rate m", "AMAT_I [Takte]", "AMAT_D [Takte]",
              "I$_stall/Instr", "D$_stall/Instr", "Mem_Stall_CPI",
              "CPI_eff", "IPC_eff", "Speed-Up vs. ideal", "Zeitverlust [%]"]

for t_miss, scenario_label in MISS_PENALTIES:
    # Scenario header
    s(ws2, r, 1, f"Szenario: {scenario_label}  |  t_miss = {t_miss} Systemtakte",
      FILL_DARK, fwh(sz=11), ha="left", cs=13)
    ws2.row_dimensions[r].height = 20; r+=1

    # Column headers
    for ci, lbl_txt in enumerate(COL_LABELS, start=1):
        shdr(ws2, r, ci, lbl_txt)
    shdr(ws2, r, 11, "CPI-Diagramm (normiert)")
    shdr(ws2, r, 12, "Klassifikation")
    ws2.row_dimensions[r].height = 30; r+=1

    for i, m in enumerate(MISS_RATES):
        amat_i   = 1 + m * t_miss
        amat_d   = 1 + m * t_miss
        i_stall  = m * t_miss
        d_stall  = 0.25 * m * t_miss        # 25 % load frequency
        mem_stall = i_stall + d_stall
        cpi_eff  = 1 + mem_stall
        ipc      = 1.0 / cpi_eff
        speedup  = cpi_eff                  # vs ideal CPI=1, SpeedUp = CPI_eff
        loss_pct = (cpi_eff - 1) * 100

        if m <= 0.01:    bg = FILL_GREEN
        elif m <= 0.05:  bg = FILL_BGRN
        elif m <= 0.10:  bg = FILL_YELLOW
        elif m <= 0.20:  bg = FILL_ORANGE
        else:            bg = FILL_RED

        cls = ("sehr gut"     if m <= 0.01 else
               "gut"          if m <= 0.05 else
               "akzeptabel"   if m <= 0.10 else
               "problematisch"if m <= 0.20 else
               "kritisch")

        vals = [f"{m*100:.1f} %", amat_i, amat_d, i_stall, d_stall,
                mem_stall, cpi_eff, ipc, speedup, loss_pct]
        nfmts = ["@", "0.00", "0.00", "0.00", "0.00",
                 "0.00", "0.00", "0.000", "0.00", "0.0"]
        for ci, (v, nf_) in enumerate(zip(vals, nfmts), start=1):
            num(ws2, r, ci, v, nf=nf_, f=bg)
        # Mini-bar (text)
        bar_len = max(1, int(mem_stall))
        bar = "█" * min(bar_len, 20) + ("+" if bar_len > 20 else "")
        s(ws2, r, 11, bar, f=bg, fnt=Font(color="1F4E79"))
        s(ws2, r, 12, cls, f=bg, fnt=fbk(bold=(m > 0.1)))
        r+=1

    # Legend for this block
    s(ws2, r, 1, "Grün ≤ 1 % | Hell-Grün ≤ 5 % | Gelb ≤ 10 % | Orange ≤ 20 % | Rot > 20 %",
      FILL_GREY, fbk(bold=False, sz=9), cs=13)
    r+=2

# ===========================================================================
# SHEET 3 – Takt-Matrix
# ===========================================================================
ws3 = wb.create_sheet("Takt_Matrix")
col_widths = [5, 18, 14, 14, 14, 18, 18, 18, 18, 14, 18, 18, 18, 26]
for ci, w in enumerate(col_widths, start=1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

r = 1
s(ws3, r, 1, "Takt-Kombinations-Matrix  —  RWU-RV64I  (QSPI / CPU-Core / SoC / Eingangs-Takt)",
  FILL_DARK, fwh(sz=12), ha="center", cs=14)
ws3.row_dimensions[r].height = 24; r+=2

# Flash device info box
shdr(ws3, r, 1, "Referenz-Flash-Geräte", cs=14); r+=1
flash_rows = [
    ("Winbond W25Q128JV",  "128 Mbit (16 MB)", "133 MHz",  "80 MHz (konservativ, 2.7–3.6 V)",
     "0x6B (Quad Out, 1-1-4)", "8", "0xEF", "★ Primär-Referenz laut Spec"),
    ("Micron MT25QL128",   "128 Mbit (16 MB)", "133 MHz",  "80 MHz (konservativ, 2.7–3.6 V)",
     "0x6B (Quad Out, 1-1-4)", "8", "0x20", "★ Sekundär-Referenz laut Spec"),
]
fhdr_cols = ["Gerät", "Kapazität", "Max SCK (Datenblatt)", "Max SCK (Design-Ziel)",
             "Quad-Read Opcode", "Dummy-Zyklen", "JEDEC-ID", "Anmerkung"]
for ci, h in enumerate(fhdr_cols, start=1):
    shdr(ws3, r, ci+1 if ci > 7 else ci,  # shift hack avoided below
         h if ci <= len(fhdr_cols) else "")
# rebuild properly:
ws3.row_dimensions[r].height = 0  # hide the broken row
r+=1
for ci, h in enumerate(fhdr_cols, start=1):
    shdr(ws3, r, ci, h)
r+=1
for i, fr in enumerate(flash_rows):
    bg = FILL_GREEN if i == 0 else FILL_LIGHT
    for ci, v in enumerate(fr, start=1):
        s(ws3, r, ci, v, f=bg, fnt=fbk(bold=(ci==1 or ci==8)), wrap=True)
    r+=1
r+=1

# --- Clock Matrix Headers ---------------------------------------------------
COL_HDRS = [
    "#", "Quarz f_in\n[MHz]", "PLL\n(ja/nein)", "PLL-Faktor\nN",
    "f_sys / f_CPU\n[MHz]", "QSPI-Takt\nQuelle", "QSPI-Divider\n(÷)",
    "f_QSPI_SCK\n[MHz]", "t_QSPI-Takt\n[ns]", "QSPI-Zyklen\n(Protokoll)",
    "t_miss\n[µs]", "Miss-Penalty\n[Systemtakte]",
    "W25Q128JV\n(≤80 MHz?)", "MT25QL128\n(≤80 MHz?)",
]
for ci, h in enumerate(COL_HDRS, start=1):
    c = shdr(ws3, r, ci, h)
    ws3.row_dimensions[r].height = 40
    ws3.cell(row=r, column=ci).alignment = Alignment(horizontal="center",
                                                      vertical="center",
                                                      wrap_text=True)
r+=1

# Second header row for AMAT at different miss rates + recommendation
AMAT_MR = [0.01, 0.05, 0.10, 0.20]
amat_start_col = len(COL_HDRS) + 1
for ci, mr in enumerate(AMAT_MR, start=amat_start_col):
    shdr(ws3, r-1, ci, f"AMAT\n@ {int(mr*100)} % Miss")
shdr(ws3, r-1, amat_start_col + len(AMAT_MR), "Timing-Risiko\nMISO-Setup")
shdr(ws3, r-1, amat_start_col + len(AMAT_MR)+1, "Empfehlung")
ws3.column_dimensions[get_column_letter(amat_start_col)].width   = 14
ws3.column_dimensions[get_column_letter(amat_start_col+1)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+2)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+3)].width = 14
ws3.column_dimensions[get_column_letter(amat_start_col+4)].width = 16
ws3.column_dimensions[get_column_letter(amat_start_col+5)].width = 28

QSPI_PROTO_CYCLES = 106   # from spec: 8+24+8+64+2

# Clock combinations to enumerate
# (f_in, pll_en, pll_n, f_sys, qspi_source, qspi_div)
# qspi_source: "=f_sys/div" or "=f_in direkt"
combos = []
for f_in in [10, 20, 25]:
    for pll_n in [1, 2, 4, 5]:
        f_sys = f_in * pll_n
        if f_sys > 60 or f_sys < 5:
            continue
        pll_en = "nein" if pll_n == 1 else "ja"
        for qdiv in [1, 2, 4]:
            f_qspi = f_sys / qdiv
            combos.append((f_in, pll_en, pll_n, f_sys, "f_sys / Div.", qdiv, f_qspi))

# Also add "QSPI = f_in direct" (crystal drives QSPI, separate path)
for f_in in [10, 25]:
    for pll_n in [1, 2, 5]:
        f_sys = f_in * pll_n
        if f_sys > 60 or f_sys < 5: continue
        pll_en = "nein" if pll_n == 1 else "ja"
        f_qspi = f_in   # crystal direct
        if f_qspi != f_sys:   # skip if same as div=1 already covered
            combos.append((f_in, pll_en, pll_n, f_sys, "f_in (direkt)", 1, f_qspi))

# Sort by f_sys desc, then f_qspi desc
combos.sort(key=lambda x: (-x[3], -x[6]))

# De-duplicate
seen = set()
unique_combos = []
for c in combos:
    key = (c[0], c[3], c[4], c[6])
    if key not in seen:
        seen.add(key); unique_combos.append(c)

# "Best" combinations for both flash devices:
# Criteria: f_qspi ≤ 80 MHz, MISO-timing safe, miss_penalty minimised, f_sys maximised
# We mark: ★★ best, ★ good
def timing_risk(f_sys, f_qspi, qspi_src):
    """MISO setup-time risk for synchronous QSPI controller."""
    if qspi_src == "f_in (direkt)":
        return "gering (async. Taktdomäne)"
    # tCO_max(W25Q128JV) ≈ 7 ns + ~3 ns PCB  → 10 ns total
    # Need: (1/f_qspi) - 10ns > setup_margin (5 ns min)
    period_ns = 1000.0 / f_qspi
    margin_ns = period_ns - 10          # generous: tCO=7ns + PCB=3ns
    if margin_ns >= 15:
        return "gering"
    elif margin_ns >= 8:
        return "moderat"
    elif margin_ns >= 2:
        return "erhöht – Layout prüfen"
    else:
        return "kritisch – nicht empfohlen"

def recommend(f_in, pll_en, pll_n, f_sys, qspi_src, f_qspi, t_miss_us):
    risk = timing_risk(f_sys, f_qspi, qspi_src)
    ok_flash = f_qspi <= 80
    if not ok_flash:
        return ("", "Außerhalb Flash-Spec")
    if "kritisch" in risk:
        return ("", "Timing-Risiko")
    # Penalize high dividers (= high miss penalty)
    miss_penalty = QSPI_PROTO_CYCLES * f_sys / f_qspi
    if f_sys >= 50 and f_qspi >= 25 and miss_penalty <= 212 and "erhöht" not in risk:
        if miss_penalty == 106 and f_sys == 50:
            return ("★★ BESTE OPTION", f"f_sys={f_sys} MHz, SCK={f_qspi} MHz, t_miss=106 Takte")
        if miss_penalty == 212 and f_sys == 50:
            return ("★★ EMPFOHLEN", f"f_sys={f_sys} MHz, f_QSPI={f_qspi} MHz, Timing sicher")
    pll_note = "kein PLL" if pll_en == "nein" else f"PLL×{pll_n}"
    if f_sys >= 25 and f_qspi >= 25 and miss_penalty <= 106 and "erhöht" not in risk:
        return ("★ GUT", f"{pll_note}, f_sys={f_sys} MHz, f_QSPI={f_qspi} MHz")
    if f_sys >= 25 and f_qspi >= 25:
        return ("◎ akzeptabel", "")
    return ("", "")

for idx, (f_in, pll_en, pll_n, f_sys, qspi_src, qdiv, f_qspi) in enumerate(unique_combos, start=1):
    t_qspi_ns   = 1000.0 / f_qspi if f_qspi > 0 else 0
    t_miss_us   = QSPI_PROTO_CYCLES / f_qspi if f_qspi > 0 else 0
    miss_pen    = QSPI_PROTO_CYCLES * f_sys / f_qspi if f_qspi > 0 else 0
    ok_w        = "✓" if f_qspi <= 80 else "✗  (zu schnell)"
    ok_m        = "✓" if f_qspi <= 80 else "✗  (zu schnell)"
    risk_str    = timing_risk(f_sys, f_qspi, qspi_src)
    rec_sym, rec_note = recommend(f_in, pll_en, pll_n, f_sys, qspi_src, f_qspi, t_miss_us)

    # Row background
    if "★★" in rec_sym:
        bg = FILL_BGRN
    elif "★" in rec_sym:
        bg = FILL_GREEN
    elif f_qspi > 80:
        bg = FILL_RED
    elif "kritisch" in risk_str:
        bg = FILL_ORANGE
    elif idx % 2 == 0:
        bg = FILL_GREY
    else:
        bg = FILL_WHITE

    row_vals = [
        idx, f_in,
        pll_en, (pll_n if pll_en == "ja" else "—"),
        f_sys, qspi_src, qdiv,
        f_qspi, f"{t_qspi_ns:.1f}", QSPI_PROTO_CYCLES,
        f"{t_miss_us:.2f}", f"{miss_pen:.0f}",
        ok_w, ok_m,
    ]
    for ci, v in enumerate(row_vals, start=1):
        ha = "center" if ci not in (6,) else "left"
        fnt = fbk(bold=("★" in rec_sym))
        s(ws3, r, ci, v, f=bg, fnt=fnt, ha=ha)
    # AMAT columns
    for ci2, mr in enumerate(AMAT_MR, start=amat_start_col):
        amat_val = 1 + mr * miss_pen
        num(ws3, r, ci2, round(amat_val, 1), nf="0.0", f=bg)
    # Timing risk
    risk_bg = (FILL_GREEN if "gering" in risk_str else
               FILL_YELLOW if "moderat" in risk_str else
               FILL_ORANGE if "erhöht" in risk_str else
               FILL_RED)
    s(ws3, r, amat_start_col+len(AMAT_MR), risk_str, f=risk_bg, fnt=fbk(), wrap=True)
    # Recommendation
    rec_bg = FILL_BGRN if "★★" in rec_sym else (FILL_GREEN if "★" in rec_sym else bg)
    rec_text = rec_sym + (f"\n{rec_note}" if rec_note else "")
    s(ws3, r, amat_start_col+len(AMAT_MR)+1, rec_text,
      f=rec_bg, fnt=fbk(bold=("★" in rec_sym)), wrap=True)
    ws3.row_dimensions[r].height = 18
    r+=1

r+=1
# Legend
legend_rows = [
    (FILL_BGRN,   "★★ BESTE OPTION / EMPFOHLEN  – optimales Verhältnis aus Performance und Timing-Sicherheit"),
    (FILL_GREEN,  "★ GUT  – geeignet für Referenzdesign"),
    (FILL_GREY,   "◎ akzeptabel – funktioniert, aber suboptimale Miss-Penalty oder Performance"),
    (FILL_ORANGE, "Erhöhtes Timing-Risiko – MISO-Setup muss im Place&Route verifiziert werden"),
    (FILL_RED,    "Nicht empfohlen – außerhalb Flash-Spec oder kritisches Timing"),
]
shdr(ws3, r, 1, "Farb-Legende", cs=amat_start_col+len(AMAT_MR)+1); r+=1
for bg, txt in legend_rows:
    s(ws3, r, 1, txt, f=bg, fnt=fbk(), cs=amat_start_col+len(AMAT_MR)+1)
    r+=1

# Freeze panes
ws3.freeze_panes = ws3.cell(row=7, column=2)

# ===========================================================================
# Save
# ===========================================================================
wb.save(OUT)
print(f"Saved: {OUT}")
